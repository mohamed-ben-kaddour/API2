from flask import Flask, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from supabase import create_client, Client
import os

# Initialize Flask app
app = Flask(__name__)

# Initialize Supabase client using environment variables
url = os.getenv("SUPABASE_URL")
key = os.getenv("SUPABASE_KEY")
supabase: Client = create_client(url, key)

@app.route('/download_excel')
def download_excel():
    try:
        response = supabase.rpc("get_monthly_attendance_counts").execute()
        print(response)
        if response.data:
            for row in response.data:
                print(row)
        else:
            print("No data returned.")
        
        data = response.data

        # Create an in-memory Excel file using openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Attendance"

        # Add column headers with formatting
        headers = ["Activity ID", "Month", "Male Count", "Female Count"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center headers

        # Add data rows
        row_num = 2  # Start from the second row
        for row in data:
            ws.cell(row=row_num, column=1, value=row.get("activity_id", "N/A")).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=2, value=row.get("month", "N/A")).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=3, value=row.get("male_count", "N/A")).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=4, value=row.get("female_count", "N/A")).alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1  # Move to the next row

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Save to an in-memory file
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name='rapport_presences_activite.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
